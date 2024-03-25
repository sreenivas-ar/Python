import openpyxl
from openpyxl import workbook
wb=openpyxl.load_workbook('D:\HandsOn\python\Automation\excel\ProjectManagement.xlsx')
sheet = wb['Project Management Data']
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['A1'].value)
# print(sheet.cell(row=1, column=7).value)

# print(sheet.cell(row=sheet.max_row, column=sheet.max_column).value)
for rowVal in range(1,sheet.max_row+1):
    for colVal in range(1,sheet.max_column+1):
        print(sheet.cell(row=rowVal, column=colVal).value, end=" ")
    print()