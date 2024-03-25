#pip install openpyxl
import openpyxl,os
from openpyxl import Workbook
# wb=Workbook()
# ws1 = wb.create_sheet(title='Sheet 1')

wb=openpyxl.load_workbook('D:\HandsOn\python\Automation\excel\Sample.xlsx')
print(type(wb))
print(os.getcwd())
# os.chdir()

#getting sheets from the workbook
print(wb.sheetnames) #prints sheetnames in list format
sheet=wb[wb.sheetnames[0]] #accessing first sheet by its name
sheet=wb['Sheet1']     #accessing the first sheet by its name
print(type(sheet))    #returns
print(sheet.title)
print(wb.active)

#getting cells from the sheets
print(wb['Sheet1']['A1'].value)   #accessing value of cell A1 in Sheet1

print(sheet.cell(row=1, column=2)) #accessing cell B1 using row and column
print(sheet.cell(row=1, column=2).value) #accessing value of cell B1


wb=openpyxl.load_workbook('D:\HandsOn\python\Automation\excel\Sample.xlsx')
sheet=wb['Sheet1']
print(sheet.max_row)
print(sheet.max_column)

from openpyxl.utils import get_column_letter, column_index_from_string
wb=openpyxl.load_workbook("D:/HandsOn/python/Automation/excel/Sample.xlsx")
print(get_column_letter(3))       #converts column number to letter
print(column_index_from_string('AA'))      #convert letter to column number

print(tuple(wb['Sheet1']['A1':'C1']))   #read a range of cells as tuple
for rowRe in wb['Sheet1']['A1': 'B2']:
    for colRe in rowRe:
        print(colRe.coordinate, colRe.value)
    print('----------------------------')

wb=openpyxl.load_workbook('D:\HandsOn\python\Automation\excel\Sample.xlsx')
print(wb.active)
sheet=wb.active
print(list(sheet.columns)[1])    #accessing data in second column
for colVal in list(sheet.columns)[1]:
    print(colVal.value)         #prints all values in 2nd column